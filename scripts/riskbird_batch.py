# Maintained by Lu Lingyan, Deheng (Wuxi) Law Firm. / Adapted for agent-browser by WorkBuddy
#!/usr/bin/env python3
"""批量/单查通过风鸟(riskbird.com)查询企业电话、邮箱及工商信息（agent-browser 适配版）

适配说明：原版依赖 CDP Proxy(localhost:3456) 操作已登录 Chrome，
本版改用 agent-browser CLI 操作浏览器（风鸟不拦截 agent-browser，已实测）。

用法：
  # 文件/表格输入（自动识别 .xlsx/.csv/.json/.txt 扩展名，多工作表自动合并去重）
    python riskbird_batch.py 名单.xlsx
    python riskbird_batch.py 名单.csv
    python riskbird_batch.py 名单.json
    python riskbird_batch.py 名单.txt                  # 每行一个公司名
    python riskbird_batch.py 名单.xlsx --out 结果.xlsx
    python riskbird_batch.py 名单.xlsx --dry           # 仅抽取公司名，不查风鸟（校验用）

  # 直接文字输入公司名（无需准备文件）
  #  · 单家：直接在对话流打印结果，不生成表格
  #  · 批量：--name/--names 传多个，自动输出 Excel 表格
    python riskbird_batch.py --name "北京博维伟业有限公司"                   # 单家搜索（对话流告知）
    python riskbird_batch.py --name "A公司" --name "B公司" --name "C公司"   # 批量搜索（输出表格）
    python riskbird_batch.py --names "A公司" --names "B公司"                # 同上，等价写法

Excel 多表处理：逐个工作表自动识别“公司名”列（按含 公司/科技/电子/贸易… 正则挑选，
再过滤标题/表头行），跨表按公司名去重。

输出：
  · 文件/批量输入 → 默认生成 同名_联系方式.xlsx（字段见 EXCEL_COLS）+ 同名_联系方式.json（断点续跑）。
  · 单家文字输入 → 直接在对话流打印电话/邮箱/法人等，不落表；如需落表可加 --out 路径。
  · 自定义输出：--out 指定 Excel 路径；--json-out 指定 JSON 明细路径。

前置：agent-browser 已安装且用固定 profile 登录过风鸟（登录态会保留在 agent-browser session）。
"""
import argparse
import json
import subprocess
import time
import os
import re
import urllib.parse
import sys

DEBUG = True

_AGENT_BROWSER = os.environ.get(
    "AGENT_BROWSER_EXE",
    r"C:\Users\Administrator\.workbuddy\binaries\node\versions\22.22.2\node_modules\agent-browser\bin\agent-browser-win32-x64.exe",
)

# 固定登录态目录：首次登录后，后续运行复用同一会话（登录态持久化）
RISKBIRD_PROFILE = os.environ.get(
    "RISKBIRD_PROFILE",
    os.path.expanduser(r"~\.workbuddy\riskbird-profile"),
)

EXCEL_COLS = ["序号", "企业名", "注册地点",  "法人", "电话", "邮箱", "规模", "注册资本", "来源表", "经营范围"]

NAME_HEADERS = ("企业名", "企业名称", "公司名", "公司名称", "名称", "公司", "企业", "客户", "代理", "商户",
                "company", "name")

# 判定“像公司名”的正则：含这些词且非空，用于挑列 + 过滤标题行
COMPANY_RE = re.compile(r"(公司|集团|企业|科技|电子|网络|贸易|信息|实业|商贸|股份|有限责任|研究|中心|厂|店|商行|有限)")


def _clean_env():
    env = os.environ.copy()
    for k in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"):
        env.pop(k, None)
    env["NO_PROXY"] = "*"
    env["no_proxy"] = "*"
    return env


def _ab(args, timeout=60):
    cmd = [_AGENT_BROWSER, "--profile", RISKBIRD_PROFILE] + args
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, env=_clean_env())
    return r.stdout.strip()


def navigate(url):
    _ab(["open", url])


def get_page_text(max_len=10000):
    out = _ab(["eval", f"document.body.innerText.slice(0,{max_len})"])
    try:
        val = json.loads(out)
        return val if isinstance(val, str) else out
    except Exception:
        return out


def extract_from_search_results(text, company_name):
    lines = text.split("\n")
    info = {"company": company_name, "phone": "", "email": "", "website": "", "socialSecurity": "",
            "staffSize": "", "address": "", "legalPerson": "", "capital": "", "establishDate": "",
            "status": "", "creditCode": ""}
    target_idx = -1
    for i, line in enumerate(lines):
        if line.strip() == company_name:
            target_idx = i
            break
    if target_idx < 0:
        for i, line in enumerate(lines):
            if company_name in line and len(line.strip()) > 5:
                target_idx = i
                break
    if target_idx < 0:
        return None
    for j in range(target_idx, min(target_idx + 30, len(lines))):
        line = lines[j].strip()
        if not line:
            continue
        if line in ("在营", "正常", "存续", "在营/正常", "吊销", "注销"):
            info["status"] = line
        if line.startswith("电话：") or line.startswith("电话:"):
            val = line.split("：")[-1].split(":")[-1].strip()
            m = re.search(r'(\d[\d\-]{5,})', val)
            if m:
                info["phone"] = m.group(1)
        if line.startswith("邮箱：") or line.startswith("邮箱:"):
            val = line.split("：")[-1].split(":")[-1].strip()
            m = re.search(r'([A-Za-z0-9._%-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})', val)
            if m:
                info["email"] = m.group(1)
        if line.startswith("官网：") or line.startswith("官网:") or line.startswith("网址："):
            val = line.split("：")[-1].split(":")[-1].strip()
            if val and val != "-":
                info["website"] = val
        if "通信地址" in line or "注册地址" in line:
            parts = re.split(r'[：:]', line)
            if len(parts) >= 2:
                val = parts[-1].strip()
                if val and val != "-":
                    info["address"] = val
        if "法定代表人" in line and "：" in line:
            info["legalPerson"] = line.split("：")[-1].strip()
        elif "法定代表人" in line and j + 1 < len(lines):
            next_l = lines[j + 1].strip()
            if next_l and not next_l.startswith("注册资本") and not next_l.startswith("电话"):
                info["legalPerson"] = next_l.split("：")[-1].strip()
        if "注册资本" in line and "实缴" not in line and "：" in line:
            info["capital"] = line.split("：")[-1].strip()
        if "成立日期" in line and "：" in line:
            info["establishDate"] = line.split("：")[-1].strip()
        if "统一社会信用代码" in line and "：" in line:
            info["creditCode"] = line.split("：")[-1].strip()
    if not info["phone"]:
        context = "\n".join(lines[max(0, target_idx - 2):target_idx + 20])
        phone_match = re.search(r'电话[：:]\s*(\d[\d\-]{6,})', context)
        if phone_match:
            info["phone"] = phone_match.group(1).strip()
    if not info["email"]:
        context = "\n".join(lines[max(0, target_idx - 2):target_idx + 20])
        email_match = re.search(r'邮箱[：:]\s*([\w.%-]+@[\w.-]+\.[\w]{2,})', context)
        if email_match:
            info["email"] = email_match.group(1).strip()
    return info


def query_one(company):
    encoded = urllib.parse.quote(company)
    search_url = f"https://www.riskbird.com/search/company?keyword={encoded}&_t={int(time.time() * 1000)}"
    navigate(search_url)
    time.sleep(6)
    text = get_page_text(10000)
    if "额度已用完" in text or "达到上限" in text or "去登录" in text or "登录/注册" in text:
        return {"company": company, "error": "NEED_LOGIN",
                "message": "风鸟未登录或会话已过期，请先用 agent-browser 登录一次（--profile 固定目录），再重跑"}
    info = extract_from_search_results(text, company)
    if info is None:
        return {"company": company, "error": "NOT_FOUND", "_preview": text[:500]}
    info["query_time"] = time.strftime("%Y-%m-%d %H:%M:%S")
    return info


def _looks_like_company(s):
    return bool(COMPANY_RE.search(s)) and len(s) >= 4


def _pick_company_col(rows):
    """在多个工作表/多列中挑选最可能“公司名”的列（按命中公司正则的比例）。"""
    if not rows:
        return 0
    n_cols = max((len(r) for r in rows if r), default=1)
    best, best_score = 0, -1
    for j in range(n_cols):
        score = 0
        for r in rows[:60]:
            if j < len(r) and r[j] is not None:
                v = str(r[j]).strip()
                if _looks_like_company(v):
                    score += 1
        if score > best_score:
            best, best_score = j, score
    return best


def read_companies(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".json":
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return [str(x).strip() for x in data if str(x).strip()]
        return [str(v).strip() for v in data.values() if str(v).strip()]
    if ext == ".csv":
        import csv
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        idx = 0
        if rows:
            for h in rows[0]:
                if str(h).strip().lower() in [x.lower() for x in NAME_HEADERS]:
                    idx = rows[0].index(h)
                    break
        return [r[idx].strip() for r in rows[1:] if len(r) > idx and r[idx].strip()]
    if ext == ".txt":
        with open(path, "r", encoding="utf-8") as f:
            return [l.strip() for l in f if l.strip()]
    if ext in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        seen = {}
        order = []
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            col = _pick_company_col(rows)
            for r in rows:
                if col < len(r) and r[col] is not None:
                    v = str(r[col]).strip()
                    if _looks_like_company(v) and v not in seen:
                        seen[v] = ws.title
                        order.append((v, ws.title))
        return order  # list of (name, source_sheet)
    raise ValueError(f"不支持的输入格式: {ext}（支持 .json/.csv/.txt/.xlsx）")


def write_excel(results, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.append(EXCEL_COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(results, 1):
        ws.append([
            i, r.get("company", ""),
            r.get("address", ""), r.get("legalPerson", ""),
            r.get("phone", ""), r.get("email", ""),
            r.get("staffSize", ""), r.get("capital", ""),
            r.get("source", ""), "",
        ])
    widths = [6, 40, 30, 16, 18, 28, 10, 16, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(path)


def save_results(results, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)


def _format_single(info):
    """单家查询结果格式化为对话流文本（无表格）。"""
    if info.get("error"):
        err = info.get("error")
        if err == "NEED_LOGIN":
            msg = info.get("message") or "风鸟未登录或会话已过期"
            tip = "\n请先用 agent-browser 以固定 profile 登录风鸟后再重跑。"
        elif err == "NOT_FOUND":
            msg = "风鸟未匹配到该企业"
            tip = "\n请确认公司名是否准确（需与工商全称一致）。"
        else:
            msg = info.get("message") or err
            tip = ""
        return f"⛔ 查询失败：{info.get('company','')}\n原因：{msg}{tip}"
    lines = [
        "📋 查询结果：" + info.get("company", ""),
        "─" * 48,
        f"  电话：{info.get('phone') or '—'}",
        f"  邮箱：{info.get('email') or '—'}",
        f"  法人：{info.get('legalPerson') or '—'}",
        f"  注册资本：{info.get('capital') or '—'}",
        f"  成立日期：{info.get('establishDate') or '—'}",
        f"  经营状态：{info.get('status') or '—'}",
        f"  统一信用代码：{info.get('creditCode') or '—'}",
        f"  地址：{info.get('address') or '—'}",
        "─" * 48,
    ]
    if not info.get("phone") and not info.get("email"):
        lines.append("⚠️ 未提取到电话/邮箱（页面未展示或企业未公开）")
    return "\n".join(lines)


def _run_single(args, company):
    """单家公司名查询：直接打印到对话流，不强制落表（--out 显式指定时额外落地）。"""
    print(f"🔍 查询中：{company}\n")
    info = query_one(company)
    print(_format_single(info))
    if args.out:
        write_excel([info], args.out)
        print(f"\nExcel: {args.out}")


def main():
    ap = argparse.ArgumentParser(
        description="风鸟批量/单查企业联系方式（支持文件、单名、多名）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="详见脚本顶部用法说明。",
    )
    ap.add_argument("companies_file", nargs="?", help="企业名单文件 .xlsx/.csv/.json/.txt（可选，改用 --name/--names 时不需要）")
    ap.add_argument("out_xlsx", nargs="?", default=None, help="输出 Excel 路径（兼容旧调用；也可用 --out）")
    ap.add_argument("--name", "-n", action="append", metavar="公司名", help="单个公司名，可重复多次（单独/批量搜索）")
    ap.add_argument("--names", "-N", action="append", metavar="公司名", help="同 --name，多个公司名可重复传入")
    ap.add_argument("--out", "-o", default=None, help="输出 Excel 路径")
    ap.add_argument("--json-out", default=None, help="输出 JSON 明细路径（默认与 out 同基名 _联系方式.json）")
    ap.add_argument("--dry", action="store_true", help="仅抽取/列出公司名，不查风鸟")
    args = ap.parse_args()

    # 收集公司名：优先命令行直输，其次文件
    cli_names = []
    for n in (args.name or []):
        n = n.strip()
        if n:
            cli_names.append((n, "命令行"))
    for n in (args.names or []):
        n = n.strip()
        if n:
            cli_names.append((n, "命令行"))

    src_map = {}
    if cli_names:
        names = [x[0] for x in cli_names]
        src_map = {x[0]: x[1] for x in cli_names}
    elif args.companies_file:
        if not os.path.exists(args.companies_file):
            print(f"❌ 找不到名单文件: {args.companies_file}")
            sys.exit(1)
        raw = read_companies(args.companies_file)
        if raw and isinstance(raw[0], tuple):
            names = [x[0] for x in raw]
            src_map = {x[0]: x[1] for x in raw}
        else:
            names = [str(x) for x in raw]
            src_map = {}
    else:
        ap.print_help()
        sys.exit(1)

    if not names:
        print("⚠️ 未解析到任何企业名（请检查文件表头/格式，或用 --name 直接输入）")
        sys.exit(1)

    # dry 离线校验（无论单家/批量都先处理，不查风鸟）
    if args.dry:
        from collections import Counter
        print(f"🔍 离线抽取（不查风鸟）共 {len(names)} 家公司：")
        cnt = Counter(src_map.get(n, "?") for n in names)
        for src, c in cnt.items():
            print(f"  - {src}: {c} 家")
        for n in names[:20]:
            print("   ", n)
        print(f"   ...（共 {len(names)} 家）")
        return

    # 单家文字查询：直接在对话流告知结果，不生成表格
    if len(cli_names) == 1 and not args.companies_file:
        _run_single(args, cli_names[0][0])
        return

    # 输出路径（批量/文件模式）
    out_xlsx = args.out or args.out_xlsx
    if not out_xlsx:
        if args.companies_file:
            base = os.path.splitext(args.companies_file)[0]
        else:
            base = os.path.join(os.getcwd(), "搜索结果")
        out_xlsx = base + "_联系方式.xlsx"
    _base = os.path.splitext(out_xlsx)[0]
    if _base.endswith("_联系方式"):
        _base = _base[: -len("_联系方式")]
    json_file = args.json_out or (_base + "_联系方式.json")

    print("=" * 60)
    print(f"风鸟查询 | 共 {len(names)} 家企业 | 搜索页直接提取")
    print("=" * 60)

    results = []
    if os.path.exists(json_file):
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                results = json.load(f)
            print(f"📂 已有记录: {len(results)} 条")
        except Exception:
            pass

    done = set(r.get("company", "") for r in results if r.get("company") and not r.get("error"))
    todo = [n for n in names if n not in done]
    print(f"✅ 已完成: {len(done)} | ⏳ 待查询: {len(todo)}\n")

    if not todo:
        print("全部完成！")
    else:
        success = fail = 0
        for idx, company in enumerate(todo):
            try:
                info = query_one(company)
            except Exception as e:
                info = {"company": company, "error": f"异常: {e}"}
            info["source"] = src_map.get(company, "")
            results.append(info)
            save_results(results, json_file)
            err = info.get("error", "")
            if err:
                fail += 1
                if "QUOTA" in err or "NEED_LOGIN" in err:
                    print(f"  ⛔ 额度/登录异常（{err}），停止后续查询")
                    break
            else:
                success += 1 if info.get("phone") else 0
                if not info.get("phone"):
                    fail += 1
            if (idx + 1) % 10 == 0 or idx + 1 == len(todo):
                print(f"  进度 {idx+1}/{len(todo)} | 成功 {success} | 失败 {fail}")
            time.sleep(2)
        print(f"\n{'=' * 60}")
        print(f"完成！成功: {success} | 失败: {fail} | 总计: {len(results)}")

    save_results(results, json_file)
    write_excel(results, out_xlsx)
    print(f"JSON: {json_file}")
    print(f"Excel: {out_xlsx}")


if __name__ == "__main__":
    main()
