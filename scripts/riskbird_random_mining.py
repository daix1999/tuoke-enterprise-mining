#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
风鸟随机关键词组合挖掘 —— 按头儿的设计思路整合

设计思路：
1. 关键词多维度池：经营范围 / 注册地点 / 注册资本 / 行业 / 成立年限 / 企业类型
2. 每次运行随机抽取组合（经营范围必选 + 随机叠加 0~3 个维度）
3. 串码识别组合（SHA1 前 8 位），同一组合永远同一串码
4. 台账记录：组合串码 + 运行日期时间
5. 每次只输出 50 家新企业（按企业名去重，历史输出过的不再出）
6. 表名 = 关键词 + 日期 + 串码，首行 = 关键词组合
7. 联系方式批量填充（风鸟查公司，免费）

子命令：
  combo   随机生成一个组合 + 串码 + 查台账（是否用过）
  mine    随机组合 -> 风鸟API搜索 -> 去重 -> 输出50家 -> 记录台账
  stats   台账统计

用法:
  python riskbird_random_mining.py combo --db <ledger.db>
  python riskbird_random_mining.py mine --db <ledger.db> --count 50 --outdir <输出目录>
"""
import argparse
import hashlib
import json
import os
import random
import re
import sqlite3
import subprocess
import sys
import time
from datetime import datetime

import requests

API = "https://www.riskbird.com/riskbird-api/api/v1/senior-search/search"
AB = os.environ.get(
    "AGENT_BROWSER_EXE",
    r"C:\Users\Administrator\.workbuddy\binaries\node\versions\22.22.2\node_modules\agent-browser\bin\agent-browser-win32-x64.exe",
)
ENTTYPE_LTD = "1100,2100,5100,5500,5815,5817,6100,6811,6815,7120,10400,11501,21501"
CONTACT_PHONE = "has_lianxi_phone"
STATUS_ACTIVE = "1"  # 企业状态：在营/正常（存续），过滤注销/吊销/停业

# ---------------- 关键词维度池（可自行增删） ----------------
DIMENSIONS = {
    # 注册地点（随机叠加，城市 -> 行政区划代码）
    "region": {"北京": "110000", "上海": "310000", "天津": "120000",
               "广州": "440100", "深圳": "440300", "杭州": "330100",
               "南京": "320100", "成都": "510100", "武汉": "420100",
               "西安": "610100", "重庆": "500000", "苏州": "320500",
               "郑州": "410100", "济南": "370100", "青岛": "370200",
               "长沙": "430100", "福州": "350100", "厦门": "350200",
               "合肥": "340100"},
    # 注册资本（必选，只保留中小型档，上限 1000 万过滤巨头/大公司）
    "regcap": [
        {"value": "100￥500", "scale": "小型"},
        {"value": "100￥1000", "scale": "中小型"},
        {"value": "500￥1000", "scale": "中型"},
    ],
    # 行业（随机叠加，行业名 -> 国标代码）
    "nicid": {"软件和信息技术服务业": "I65", "批发业": "F51",
              "互联网和相关服务": "I64", "电信": "I63"},
    # 成立年限（随机叠加，年限名 -> esdate 格式"上限￥下限"）
    "esdate": {"1-3年": "3￥1", "3-5年": "5￥3", "5-10年": "10￥5"},
}
MAX_FILTERS = 3  # 随机叠加维度上限

# ---------------- 身份 → 经营范围关键词映射 ----------------

IDENTITY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "industry_identities.json")


def load_identities():
    with open(IDENTITY_FILE, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("identities", {}), data.get("default", ""), data.get("customer_types", [])


IDENTITIES, DEFAULT_IDENTITY, CUSTOMER_TYPES = load_identities()

# ---------------- 工具函数 ----------------

def clean_env():
    env = os.environ.copy()
    for k in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy",
              "ALL_PROXY", "all_proxy"):
        env.pop(k, None)
    env["NO_PROXY"] = "*"
    env["no_proxy"] = "*"
    return env


def get_cookie():
    try:
        r = subprocess.run([AB, "eval", "document.cookie"], capture_output=True,
                           text=True, timeout=30, env=clean_env())
        val = json.loads(r.stdout.strip())
        return val if isinstance(val, str) else r.stdout.strip()
    except Exception:
        return ""


def combo_hash(text):
    return hashlib.sha1(text.encode("utf-8")).hexdigest()[:8]


def normalize_combo(scope, region=None, capital=None, industry=None, age=None):
    parts = []
    for k, v in [("经营范围", scope), ("地点", region), ("注册资本", capital),
                 ("行业", industry), ("成立年限", age)]:
        if v:
            parts.append(f"{k}={v}")
    return "；".join(parts)


def infer_scale(capital_text):
    if not capital_text:
        return ""
    m = re.search(r"(\d+(?:\.\d+)?)\s*万", capital_text)
    if not m:
        return ""
    try:
        wan = float(m.group(1))
    except ValueError:
        return ""
    if wan < 100:
        return "微型"
    if wan < 500:
        return "小型"
    if wan < 1000:
        return "中型"
    if wan < 5000:
        return "大型"
    return "特大型"


def random_combo(identity=None, customer_type=None):
    """根据身份(卖什么)+目标客户类型(卖给谁)取经营范围关键词 + 注册资本必选 + 行业必选 + 随机叠加地点/年限"""
    # 身份 → 客户类型 → 核心词(必带) + 叠加词池
    ident = IDENTITIES.get(identity) or IDENTITIES.get(DEFAULT_IDENTITY) or {}
    ct = customer_type or (CUSTOMER_TYPES[0] if CUSTOMER_TYPES else "")
    kw = ident.get(ct) or {"core": "计算机销售", "extra": []}
    core = kw.get("core", "计算机销售")
    extra_pool = kw.get("extra", [])
    # 经营范围：核心词必带 + 随机叠加 0~2 个叠加词（加号=同时包含）
    n_extra = random.randint(0, min(2, len(extra_pool)))
    extra = random.sample(extra_pool, k=n_extra) if extra_pool else []
    scope = "+".join([core] + extra)

    # 必选：注册资本（中小型档，过滤巨头）
    cs = random.choice(DIMENSIONS["regcap"])
    capital = cs["value"]
    capital_scale = cs["scale"]

    # 必选：行业（IT 相关，锁定渠道商/集成商）
    industry = random.choice(list(DIMENSIONS["nicid"].keys()))

    # 随机叠加：地区、成立年限（总叠加维度 ≤ 2，叠加词多时少叠地点年限，避免过窄）
    budget = max(0, 2 - n_extra)
    n = random.randint(0, budget)
    pool = ["region", "esdate"]
    chosen = random.sample(pool, k=min(n, len(pool)))
    region = age = None
    for f in chosen:
        if f == "region":
            region = random.choice(list(DIMENSIONS["region"].keys()))
        elif f == "esdate":
            age = random.choice(list(DIMENSIONS["esdate"].keys()))

    combo_text = normalize_combo(scope, region, capital, industry, age)
    return {
        "scope": scope,
        "region": region,
        "regionid": DIMENSIONS["region"].get(region, ""),
        "capital": capital,
        "scale": capital_scale,
        "industry": industry,
        "nicid": DIMENSIONS["nicid"].get(industry, ""),
        "age": age,
        "esdate": DIMENSIONS["esdate"].get(age, ""),
        "combo_text": combo_text,
        "run_id": combo_hash(combo_text),
    }


# ---------------- 台账 ----------------

SCHEMA = """
CREATE TABLE IF NOT EXISTS combos (
    run_id TEXT PRIMARY KEY, combo_text TEXT, scope TEXT, region TEXT,
    capital TEXT, scale TEXT, industry TEXT, age TEXT,
    first_date TEXT, first_time TEXT
);
CREATE TABLE IF NOT EXISTS runs (
    id INTEGER PRIMARY KEY AUTOINCREMENT, run_id TEXT,
    run_date TEXT, run_time TEXT, output_count INTEGER DEFAULT 0, output_file TEXT
);
CREATE TABLE IF NOT EXISTS companies (
    id INTEGER PRIMARY KEY AUTOINCREMENT, company_name TEXT UNIQUE,
    region TEXT, legal_person TEXT, capital TEXT, scale TEXT, business TEXT,
    phone TEXT, email TEXT, first_run_id TEXT, first_date TEXT, created_at TEXT
);
"""


def get_db(path):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.executescript(SCHEMA)
    # 兼容旧表结构（旧版 combos 无 industry 列）
    cols = [r[1] for r in conn.execute("PRAGMA table_info(combos)").fetchall()]
    if "industry" not in cols:
        conn.execute("ALTER TABLE combos ADD COLUMN industry TEXT")
    conn.commit()
    return conn


# ---------------- 风鸟 API ----------------

def build_cond(combo, extra=None):
    cond = {
        "contact": CONTACT_PHONE, "entname": "", "dom": "",
        "opscope": combo["scope"], "nicid": combo["nicid"],
        "regionid": combo["regionid"], "regcap": combo["capital"],
        "esdate": combo["esdate"], "status": STATUS_ACTIVE, "enttype": ENTTYPE_LTD,
        "orgtype": "", "ygrs": "", "sort_field": "",
        "enterprise_scale": "", "available_version": "", "tax_credit": "",
        "has_sm_ent": "", "has_jobinfo": "", "has_bid_notice": "", "has_bid_win": "",
        "has_ip_tminfo": "", "has_ip_patent": "", "has_soft_copyright": "",
        "has_work_copyright": "", "has_tuiguang_website": "", "has_icp": "",
        "has_ipr": "", "has_tuiguang_ios": "", "has_tuiguang_android": "",
    }
    if extra:
        cond.update(extra)
    return cond


def search_page(session, cond, start, length=10):
    ao_data = json.dumps([
        {"name": "sEcho", "value": start // length + 1},
        {"name": "iColumns", "value": 10},
        {"name": "sColumns", "value": "id,name,contact,email,frname,status,regCap,entType,regDate,creditNo"},
        {"name": "iDisplayStart", "value": start},
        {"name": "iDisplayLength", "value": length},
        {"name": "cSearch_conditionData", "value": cond},
    ])
    body = json.dumps({"aoData": ao_data, "queryType": "senior", "queryLimitType": 2})
    r = session.post(API, headers={"Content-Type": "application/json"}, data=body, timeout=30)
    return r.json()


# ---------------- 导出 Excel ----------------

def export_xlsx(rows, path, combo_text):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "线索"
    ws.append([combo_text])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    c0 = ws.cell(row=1, column=1)
    c0.font = Font(bold=True, size=12, color="1F4E78")
    header = ["序号", "企业名", "注册地点", "法人", "电话", "邮箱", "规模", "注册资本", "经营范围"]
    ws.append(header)
    fill = PatternFill("solid", fgColor="1F4E78")
    for col in range(1, 10):
        cell = ws.cell(row=2, column=col)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
    for i, r in enumerate(rows, 1):
        ws.append([i, r["company_name"], r["region"], r["legal_person"],
                   r.get("phone") or "", r.get("email") or "", r["scale"],
                   r["capital"], r["business"]])
    widths = [6, 40, 16, 12, 16, 24, 10, 16, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


# ---------------- 子命令 ----------------

def cmd_combo(args):
    c = random_combo(args.identity, args.customer_type)
    db = get_db(args.db)
    row = db.execute("SELECT first_date, first_time FROM combos WHERE run_id=?",
                     (c["run_id"],)).fetchone()
    db.close()
    c["used"] = row is not None
    c["history"] = {"first_date": row["first_date"], "first_time": row["first_time"]} if row else None
    print(json.dumps(c, ensure_ascii=False))


def cmd_mine(args):
    cookie = args.cookie_file and os.path.exists(args.cookie_file) and \
        open(args.cookie_file, encoding="utf-8").read().strip()
    if not cookie:
        cookie = get_cookie()
    if not cookie:
        print("❌ 无法获取登录态，请先 agent-browser 登录风鸟", file=sys.stderr)
        sys.exit(1)

    db = get_db(args.db)
    now = datetime.now()
    run_date = now.strftime("%Y-%m-%d")
    run_time = now.strftime("%H:%M:%S")

    # 随机组合（可指定 scope）。串码重复没关系，靠企业名去重保证每次输出不同
    combo = random_combo(args.identity, args.customer_type)

    # 用户自主设定的条件（必填项）覆盖随机默认值
    if args.scope:
        combo["scope"] = args.scope
    if args.region:
        combo["region"] = args.region
        combo["regionid"] = DIMENSIONS["region"].get(args.region, "")
    if args.regcap:
        combo["capital"] = args.regcap
        combo["scale"] = ""
    if args.esdate:
        combo["esdate"] = args.esdate
        combo["age"] = None
    combo["combo_text"] = normalize_combo(combo["scope"], combo["region"],
                                          combo["capital"], combo["industry"], combo["age"])
    combo["run_id"] = combo_hash(combo["combo_text"])

    # 附加筛选（不进入串码，作为搜索条件附加）
    extra = {}
    if args.contact:
        extra["contact"] = args.contact
    if args.has_bid_win:
        extra["has_bid_win"] = "1"
    if args.has_patent:
        extra["has_ip_patent"] = "1"
    if args.has_soft_copyright:
        extra["has_soft_copyright"] = "1"
    if args.has_icp:
        extra["has_icp"] = "1"

    print(f"🎲 组合: {combo['combo_text']}")
    if extra:
        print(f"   附加筛选: {extra}")

    session = requests.Session()
    session.trust_env = False
    session.headers["Cookie"] = cookie

    # 自适应放宽：命中太少时逐层去掉最窄条件（先成立年限，再地点，最后行业）
    # 用户指定的必填项不参与放宽
    relax_fields = ["esdate", "regionid", "nicid"]
    if args.esdate:
        relax_fields.remove("esdate")
    if args.region:
        relax_fields.remove("regionid")
    while True:
        cond = build_cond(combo, extra)
        data = search_page(session, cond, 0, 10)
        d = data.get("data") or {}
        if d.get("flagLimit") == "1":
            print("⚠️ 风鸟查询额度已用尽(flagLimit=1)，今日无法再搜，请明日 0 点后再跑", file=sys.stderr)
            db.close()
            sys.exit(2)
        total = d.get("totalCount", 0)
        print(f"   命中 {total} 家 (条件: {combo['combo_text']})")
        if total >= 20 or not relax_fields:
            break
        f = relax_fields.pop(0)
        if f == "esdate":
            combo["esdate"] = ""
            combo["age"] = None
        elif f == "regionid":
            combo["regionid"] = ""
            combo["region"] = None
        elif f == "nicid":
            combo["nicid"] = ""
            combo["industry"] = None
        combo["combo_text"] = normalize_combo(combo["scope"], combo["region"],
                                              combo["capital"], combo["industry"], combo["age"])
        combo["run_id"] = combo_hash(combo["combo_text"])
        print(f"   ↳ 命中太少，放宽条件重试")

    # 放宽后记录台账（最终组合）
    combo_exists = db.execute("SELECT 1 FROM combos WHERE run_id=?", (combo["run_id"],)).fetchone()
    if not combo_exists:
        db.execute("INSERT INTO combos(run_id,combo_text,scope,region,capital,scale,industry,age,first_date,first_time) "
                   "VALUES(?,?,?,?,?,?,?,?,?,?)",
                   (combo["run_id"], combo["combo_text"], combo["scope"], combo["region"],
                    combo["capital"], combo["scale"], combo["industry"], combo["age"],
                    run_date, run_time))
    print(f"   串码: {combo['run_id']} | 用过: {'是(补漏模式)' if combo_exists else '否'}")

    # 边抓边去重入库，翻页凑够 args.count 家新企业
    # 串码重复没关系：同一组合重复跑，靠企业名去重自动跳过已输出的，翻页挖新的
    cond = build_cond(combo, extra)
    added = []
    skipped = 0
    start = 0
    total = None
    max_scanned = args.count * 10  # 最多扫描 count*10 家，防重复跑时翻页过深
    while len(added) < args.count and start < max_scanned:
        data = search_page(session, cond, start, 10)
        if data.get("code") != 20000:
            break
        d = data.get("data") or {}
        if total is None:
            total = d.get("totalCount", 0)
        rows = d.get("aaData") or []
        if not rows:
            break
        for row in rows:
            name = (row.get("name") or "").strip()
            if not name:
                continue
            exists = db.execute("SELECT 1 FROM companies WHERE company_name=?", (name,)).fetchone()
            if exists:
                skipped += 1
                continue
            capital = row.get("regCap") or ""
            scale = combo["scale"] or infer_scale(capital)
            db.execute("INSERT INTO companies(company_name,region,legal_person,capital,scale,business,phone,email,first_run_id,first_date,created_at) "
                       "VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                       (name, combo["region"] or "", row.get("frname") or "", capital, scale,
                        combo["scope"], "", "", combo["run_id"], run_date,
                        now.strftime("%Y-%m-%d %H:%M:%S")))
            added.append({"company_name": name, "region": combo["region"] or "",
                          "legal_person": row.get("frname") or "", "capital": capital,
                          "scale": scale, "business": combo["scope"], "phone": ""})
            if len(added) >= args.count:
                break
        start += len(rows)
        if len(rows) < 10:
            break
        time.sleep(0.4)

    output_file = None
    if added and args.outdir:
        date_compact = run_date.replace("-", "")
        fname = f"{combo['scope']}_{date_compact}_{combo['run_id']}.xlsx"
        output_file = os.path.join(args.outdir, fname)
        export_xlsx(added, output_file, combo["combo_text"])

    db.execute("INSERT INTO runs(run_id,run_date,run_time,output_count,output_file) VALUES(?,?,?,?,?)",
               (combo["run_id"], run_date, run_time, len(added), output_file))
    db.commit()
    db.close()

    # 只打印统计，不打印名单明细（明细在 Excel 里，省 token）
    print(f"✅ 新增 {len(added)} 家 | 去重跳过 {skipped} 家 | 输出: {os.path.basename(output_file or '无')}")


def cmd_stats(args):
    db = get_db(args.db)
    n_combos = db.execute("SELECT COUNT(*) c FROM combos").fetchone()["c"]
    n_runs = db.execute("SELECT COUNT(*) c FROM runs").fetchone()["c"]
    n_companies = db.execute("SELECT COUNT(*) c FROM companies").fetchone()["c"]
    db.close()
    print(json.dumps({"combos": n_combos, "runs": n_runs, "companies": n_companies},
                     ensure_ascii=False))


def cmd_export(args):
    """按关键词组合分表导出（含电话），每次跑一轮一份独立表，不合并"""
    db = get_db(args.db)
    combos = {r["run_id"]: dict(r) for r in db.execute(
        "SELECT run_id, combo_text, scope, first_date FROM combos").fetchall()}
    rows = db.execute("SELECT company_name, region, legal_person, capital, scale, business, phone, email, first_run_id FROM companies").fetchall()
    groups = {}
    for r in rows:
        groups.setdefault(r["first_run_id"], []).append(dict(r))
    for rid, items in groups.items():
        c = combos.get(rid, {})
        scope = c.get("scope") or "未命名"
        combo_text = c.get("combo_text") or ""
        date = (c.get("first_date") or "20260817").replace("-", "")
        fname = f"{scope}_{date}_{rid}.xlsx"
        export_xlsx(items, os.path.join(args.outdir, fname), combo_text)
        print(f"导出 {fname} | {len(items)}家 | 有电话 {sum(1 for x in items if x.get('phone'))}")
    db.close()
    print(f"--- 共 {len(groups)} 份独立表（按关键词分表，不合并）")


def main():
    ap = argparse.ArgumentParser(description="风鸟随机关键词组合挖掘")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("combo", help="随机生成组合+查重")
    p.add_argument("--db", default="ledger.db")
    p.add_argument("--identity", default=None, help="用户身份(卖什么)，如'服务器/数据中心'")
    p.add_argument("--customer-type", default=None, help="目标客户类型：渠道商/经销商、集成商/项目商、终端供货")

    p = sub.add_parser("mine", help="随机组合->搜索->去重->输出50家")
    p.add_argument("--db", default="ledger.db")
    p.add_argument("--identity", default=None, help="用户身份(卖什么)，如'服务器/数据中心'")
    p.add_argument("--customer-type", default=None, help="目标客户类型：渠道商/经销商、集成商/项目商")
    p.add_argument("--scope", default=None, help="经营范围(必填项，不传则按身份随机)")
    p.add_argument("--region", default=None, help="地区(必填项)，如'北京'")
    p.add_argument("--regcap", default=None, help="注册资本区间(必填项)，如'100￥500'")
    p.add_argument("--esdate", default=None, help="成立年限(必填项)，如'5￥3'(3-5年)")
    p.add_argument("--contact", default=None, help="联系方式，如 has_lianxi_phone")
    p.add_argument("--has-bid-win", action="store_true", help="只搜有中标记录的")
    p.add_argument("--has-patent", action="store_true", help="只搜有专利的")
    p.add_argument("--has-soft-copyright", action="store_true", help="只搜有软著的")
    p.add_argument("--has-icp", action="store_true", help="只搜有备案网站的")
    p.add_argument("--count", type=int, default=50)
    p.add_argument("--outdir", default="output")
    p.add_argument("--cookie-file", default=None)

    p = sub.add_parser("stats", help="台账统计")
    p.add_argument("--db", default="ledger.db")

    p = sub.add_parser("export", help="按关键词分表导出(含电话)")
    p.add_argument("--db", default="ledger.db")
    p.add_argument("--outdir", default="output")

    args = ap.parse_args()
    {"combo": cmd_combo, "mine": cmd_mine, "stats": cmd_stats, "export": cmd_export}[args.cmd](args)


if __name__ == "__main__":
    main()
